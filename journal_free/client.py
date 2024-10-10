import cloudscraper
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from journal_free import settings


class AntiBotError(Exception):
    def __str__(self):
        return "Antibot system bypass error"


class AuthError(Exception):
    def __init__(self, msg):
        self.msg = msg

    def __str__(self):
        return f"Authentication error.\n{self.msg}"


class NZClient:
    BASE_URL = settings.BASE_URL
    session = None
    is_auth = False
    user = None
    terms: list | None = None
    selected_term = None
    journals: list | None = None
    lessons_url: list | None = None

    def __init__(self, username: str, password: str):
        self.username = username
        self.password = password
        self.session = cloudscraper.create_scraper(disableCloudflareV1=True)

    @staticmethod
    def check_antibot(response: BeautifulSoup) -> None:
        if response.find('title').text == 'Just a moment...':
            raise Exception('Antibot error, try again')

    def authenticate(self) -> bool:
        res = self.session.get(self.BASE_URL)
        if res.status_code != 200:
            if res.status_code == 403:
                raise AntiBotError
            raise Exception(f"{res.status_code}: {res.text}")
        site = BeautifulSoup(res.content, 'html.parser')
        self.check_antibot(site)
        csrf_token = site.find('input', {"name": "_csrf"}).get('value', None)
        if csrf_token is None:
            return False
        data = {
            '_csrf': csrf_token,
            'LoginForm[login]': self.username,
            'LoginForm[password]': self.password,
            'LoginForm[rememberMe]': 1,
        }
        post = self.session.post(self.BASE_URL + "/login", data=data, headers=dict(Referer=self.BASE_URL))
        post_resp = BeautifulSoup(post.content, 'html.parser')
        self.check_antibot(post_resp)
        alert = post_resp.find('div', class_='alert-danger')
        if alert is None:

            self.user = post_resp.find('div', class_='h-user-info').find('span').text
            self.is_auth = True
        else:
            self.user = None
            self.is_auth = False
            raise AuthError(alert.find('li').text)
        return self.is_auth

    def get_journals(self):
        if self.is_auth:
            res = self.session.get(self.BASE_URL + "/journal/list")
            site = BeautifulSoup(res.content, 'html.parser')
            self.check_antibot(site)
            self.terms = [{"name": term.text, "value": term['value']} for term in site.find(id="personalselectform-semester_id").find_all("option")]
            self.selected_term = {
                "name": site.find(id="personalselectform-semester_id").select_one('option[selected]').text,
                "value": site.find(id="personalselectform-semester_id").select_one('option[selected]')['value'],
            }
            journals = []
            for item in site.find("table", class_="journal-choose").find_all("tr"):
                subj_obj = item.find_all("td")
                subj_name = subj_obj[0].text
                subj_classes = subj_obj[1].find_all("a")
                journals.append({
                    "subject": subj_name,
                    "classes": [{"name": class_.text, "url": class_.get('href', None)} for class_ in subj_classes]
                })
            self.journals = journals

    def change_term(self, term_id):
        if self.is_auth:
            res = self.session.get(self.BASE_URL + "/journal/list")
            site = BeautifulSoup(res.content, 'html.parser')
            self.check_antibot(site)
            csrf_token = site.find('input', {"name": "_csrf"}).get('value', None)
            data = {
                '_csrf': csrf_token,
                'semester_id': term_id,
            }
            res = self.session.post(
                self.BASE_URL + "/site/semester-change",
                data=data,
                headers=dict(Referer=self.BASE_URL + "/journal/list")
            )
            self.get_journals()

    def find_lessons_url(self, url: str) -> None:
        res = self.session.get(self.BASE_URL + url)
        if res.status_code != 200:
            return
        site = BeautifulSoup(res.content, 'html.parser')
        self.check_antibot(site)
        pagination = site.find('ul', class_='pagination')
        page_count = len(pagination.find_all('li')) - 2 if pagination is not None else 1
        self.lessons_url = []
        for i in range(page_count):
            page_res = self.session.get(self.BASE_URL + url + '&page=' + str(i + 1))
            site = BeautifulSoup(page_res.content, 'html.parser')
            site_list_url = site.find('ul', class_='dz-container').find_all('a', class_="dz-edit modal-box")
            for site_url in site_list_url:
                self.lessons_url.append(site_url['href'])

    def add_topic(self, url: str, lesson_data: dict) -> int:
        res = self.session.get(self.BASE_URL + url)
        if res.status_code != 200:
            return res.status_code
        site = BeautifulSoup(res.content, 'html.parser')
        self.check_antibot(site)
        csrf_token = site.find('input', {"name": "_csrf"}).get('value', None)
        homework_date = site.find(id="osvitaschedulereal-hometask_to").find_all('option')[0]['value']
        data = {
            '_csrf': csrf_token,
            'OsvitaScheduleReal[lesson_topic]': lesson_data.get('topic'),
            'OsvitaScheduleReal[lesson_number_in_plan]': lesson_data.get('number'),
            'OsvitaScheduleReal[hometask]': lesson_data.get('homework'),
            'OsvitaScheduleReal[hometask_to]': homework_date,
            'OsvitaScheduleReal[second_personal_id]': '',
            'OsvitaScheduleReal[second_predmet_id]': '',
        }
        response = self.session.post(self.BASE_URL + url, data=data, headers=dict(Referer=url))
        return response.status_code

    def parse_lesson_data(self, url: str) -> dict:
        response = self.session.get(self.BASE_URL + url)
        if response.status_code != 200:
            return dict(
                status_code=response.status_code
            )
        site = BeautifulSoup(response.content, 'html.parser')
        self.check_antibot(site)
        topic = site.find(id="osvitaschedulereal-lesson_topic").get_text()
        number = site.find(id="osvitaschedulereal-lesson_number_in_plan").get('value', '')
        homework = site.find(id="osvitaschedulereal-hometask").get_text()
        return dict(
            status_code=200,
            topic=topic,
            number=number,
            homework=homework,
        )


class FileClient:
    __data: list | None = None
    valid: bool = False

    def __init__(self, file_path: str):
        document = load_workbook(file_path)
        self.active_page: Worksheet = document.active
        self._parse_data()

    def _parse_data(self) -> None:
        self.__data = []
        for row in self.active_page.rows:
            try:
                topic = row[1].value
                number = row[0].value
            except Exception:
                self.valid = False
                return

            if topic is None or number is None:
                self.valid = False
                return
            try:
                homework = row[2].value if row[2].value is not None else ""
            except IndexError:
                homework = ''
            except Exception:
                self.valid = False
                return
            self.__data.append({
                'topic': topic,
                'number': number,
                'homework': homework,
            })
        if len(self.__data) == 0:
            self.valid = False
            return
        self.valid = True

    @property
    def count(self) -> int:
        return self.active_page.max_row

    @property
    def is_valid(self) -> bool:
        return self.valid

    @property
    def validated_data(self) -> list:
        return self.__data if self.valid else []

    @property
    def _data(self) -> list:
        return self.__data

    @staticmethod
    def create(file_name: str, data: list[dict]) -> "FileClient":
        file = Workbook()
        sheet: Worksheet = file.active
        sheet.title = 'base'
        for row in data:
            sheet.append([row['number'], row['topic'], row['homework']])
        file.save(file_name)
        return FileClient(file_name)
